import time
import os
from seleniumbase import Driver
from selenium.webdriver.common.by import By
from selenium.webdriver import ActionChains
from openpyxl import Workbook, load_workbook

# Initialize the driver
driver = Driver(uc=True)
achains = ActionChains(driver)
url= input('Search any keyword in the maps and enter the url here: ')
# Navigate to the Google Maps page
driver.get(url)

# Allow time for the page to load
time.sleep(20)

# Find all relevant items
items = driver.find_elements(By.CSS_SELECTOR, 'div[role="feed"]>div>div[jsaction]')
results = []
er=0

for item in items:
    data = {
        'title': "No title available",
        'phone': "No number available",
        'web': "No website available"
    }
    try:
        # Get the title
        title_elem = item.find_element(By.CSS_SELECTOR, '.fontHeadlineSmall')
        data['title'] = title_elem.text

        # Click to reveal more information
        achains.move_to_element(title_elem).double_click().perform()
        time.sleep(2)  # Wait for the details to load

        # Get phone number
        phone_elem = driver.find_element(By.CSS_SELECTOR,
                                         'div[role="region"][aria-label$="সংক্রান্ত তথ্য"]>div>button[data-tooltip="ফোন নম্বরটি কপি করুন"]')
        data['phone'] = phone_elem.get_attribute("aria-label")

        # Get website
        web_elem = driver.find_element(By.CSS_SELECTOR,
                                       'div[role="region"][aria-label$="সংক্রান্ত তথ্য"]>div>a[data-tooltip="ওয়েবসাইটটি খুলুন"]')
        data['web'] = web_elem.get_attribute("href")

    except Exception as e:
        er+=1
        print(f"Error: {er}.No Such Element ")

    results.append(data)

# Print results
for result in results:
    print(result)


excel_file = "restaurant.xlsx"

# Check if the file exists
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
    # Remove all rows except the header
    while ws.max_row > 1:
        ws.delete_rows(2)
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Title", "Phone", "Website"])  # Header row

# Write new data to the Excel file
for result in results:
    ws.append([result['title'], result['phone'], result['web']])

# Save the workbook
wb.save(excel_file)
driver.sleep(1000)
driver.quit()